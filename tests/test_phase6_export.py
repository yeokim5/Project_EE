import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from earnings_extractor.cli import main
from earnings_extractor.export import (
    export_reviewed_run,
    format_currency_batch_text,
    format_currency_billions,
    format_gross_margin_cell,
    map_metric_to_client_cell,
)
from earnings_extractor.pipeline import extract
from earnings_extractor.review import (
    ReviewDecision,
    ReviewDecisionsFile,
    build_review_items,
    write_review_artifacts,
)
from earnings_extractor.schema import DraftRun, MetricRow
from scripts.make_acceptance_decisions import build_acceptance_decisions

ROOT = Path(__file__).resolve().parents[1]
INPUT_DIR = ROOT / "assesment_info"


def test_client_cell_mappers() -> None:
    assert format_currency_billions(22496) == "$22.5B"
    assert format_currency_billions(4100) == "$4.1B"
    assert format_currency_billions(150000) == "$150B"
    assert format_currency_batch_text(5276, "Total revenue 5,276") == "$5,276 million"
    assert (
        format_currency_batch_text(
            13900,
            "Consolidated expenses were $13.9 billion",
        )
        == "$13.9 billion"
    )
    assert format_gross_margin_cell(17.2) == 0.172

    eps = _item(_metric("Earnings per share", 0.33, unit="USD/share"))
    assert map_metric_to_client_cell(eps) == 0.33
    eps.value = 4.28
    eps.source_quote = "Diluted EPS $4.28"
    assert map_metric_to_client_cell(eps, display_mode="batch") == "$4.28 diluted"

    capital_return = _item(
        _metric(
            "Buybacks and dividends",
            (
                "$5 billion returned in 2025, including $1.6 billion of "
                "repurchases and $5.73 quarterly cash dividend per share"
            ),
            unit=None,
            scale=None,
        )
    )
    assert (
        map_metric_to_client_cell(capital_return, display_mode="batch")
        == (
            "$5.0 billion returned in 2025, including $1.6 billion of "
            "repurchases; $5.73 per share cash dividend"
        )
    )


def test_mapper_rejects_unexpected_currency_units() -> None:
    item = _item(_metric("Total revenue", 22496, unit="USD", scale="billions"))

    with pytest.raises(ValueError, match="USD millions"):
        map_metric_to_client_cell(item)


def test_demo_decisions_fail_without_override_and_pass_with_override(
    tmp_path: Path,
) -> None:
    run_dir = tmp_path / "run"
    draft_path = extract(INPUT_DIR, run_dir, mode="recorded")
    write_review_artifacts(
        run_dir,
        demo_decisions_out=run_dir / "review_decisions.json",
    )

    with pytest.raises(ValueError, match="demo-generated"):
        export_reviewed_run(
            run_dir,
            decisions_path=run_dir / "review_decisions.json",
            out_path=tmp_path / "extractions.xlsx",
        )

    artifacts = export_reviewed_run(
        run_dir,
        decisions_path=run_dir / "review_decisions.json",
        out_path=tmp_path / "extractions.xlsx",
        allow_unreviewed=True,
    )

    assert artifacts.is_draft_unreviewed is True
    workbook = load_workbook(artifacts.xlsx_path)
    assert workbook.worksheets[0].title == "Client Template DRAFT"
    assert workbook.worksheets[0]["A1"].comment is not None
    assert {"Metrics", "Review Decisions", "Evidence"}.issubset(workbook.sheetnames)
    assert DraftRun.model_validate_json(draft_path.read_text(encoding="utf-8"))


def test_acceptance_decisions_export_true_final_workbook(tmp_path: Path) -> None:
    run_dir = tmp_path / "run"
    draft_path = extract(INPUT_DIR, run_dir, mode="recorded")
    draft = DraftRun.model_validate_json(draft_path.read_text(encoding="utf-8"))
    decisions = build_acceptance_decisions(draft)
    decisions_path = run_dir / "human_review_decisions.json"
    decisions_path.write_text(decisions.model_dump_json(indent=2), encoding="utf-8")

    artifacts = export_reviewed_run(
        run_dir,
        decisions_path=decisions_path,
        out_path=tmp_path / "extractions_final.xlsx",
    )

    assert artifacts.is_draft_unreviewed is False
    workbook = load_workbook(artifacts.xlsx_path)
    worksheet = workbook.worksheets[0]
    assert worksheet.title == "Client Template"
    assert [cell.value for cell in worksheet[1]] == [
        "Company Name",
        "Quarter",
        "Total revenue",
        "Earnings per share",
        "Net income",
        "Operating income",
        "Gross margin",
        "Operating expenses",
        "Buybacks and dividends",
    ]
    assert worksheet["A2"].value == "Tesla"
    assert worksheet["C2"].value == "$22.5B"
    assert worksheet["D2"].value == 0.33
    assert worksheet["D2"].number_format == '"$"#,##0.00'
    assert worksheet["G2"].value == 0.172
    assert worksheet["G2"].number_format == "0%"
    assert worksheet["I2"].value is None
    assert worksheet["A3"].value == "Citi"
    assert worksheet["I3"].value == "$2.8B capital returned, including $1.75B buybacks"
    assert {"Metrics", "Review Decisions", "Evidence"}.issubset(workbook.sheetnames)

    evidence_headers = [cell.value for cell in workbook["Evidence"][1]]
    source_quote_col = evidence_headers.index("source_quote") + 1
    assert any(
        row[source_quote_col - 1].value
        for row in workbook["Evidence"].iter_rows(min_row=2)
    )

    payload = json.loads(artifacts.json_path.read_text(encoding="utf-8"))
    assert payload["is_draft_unreviewed"] is False
    assert artifacts.audit_report_path.name == "extractions_final.audit.md"


def test_attention_field_bare_approval_is_blocked(tmp_path: Path) -> None:
    draft = _write_draft(tmp_path)
    items = build_review_items(draft)
    decisions = _decisions_for_items(draft, items)
    attention = next(
        item for item in items if item.metric_name == "Buybacks and dividends"
    )
    for decision in decisions.decisions:
        if decision.metric_id == attention.metric_id:
            decision.review_status = "approved"
            decision.reviewer_note = None
    decisions_path = tmp_path / "decisions.json"
    decisions_path.write_text(decisions.model_dump_json(), encoding="utf-8")

    with pytest.raises(ValueError, match="requires a reviewer note"):
        export_reviewed_run(tmp_path, decisions_path, tmp_path / "out.xlsx")


def test_non_attention_rejected_blanks_cell_without_note(tmp_path: Path) -> None:
    draft = _write_draft(tmp_path)
    items = build_review_items(draft)
    decisions = _decisions_for_items(draft, items)
    net_income = next(item for item in items if item.metric_name == "Net income")
    for decision in decisions.decisions:
        if decision.metric_id == net_income.metric_id:
            decision.review_status = "rejected"
            decision.reviewer_note = None
    decisions_path = tmp_path / "decisions.json"
    decisions_path.write_text(decisions.model_dump_json(), encoding="utf-8")

    export_reviewed_run(tmp_path, decisions_path, tmp_path / "out.xlsx")

    workbook = load_workbook(tmp_path / "out.xlsx")
    assert workbook.worksheets[0]["E2"].value is None


def test_absent_template_field_blocks_final_but_exports_as_draft(
    tmp_path: Path,
) -> None:
    draft = _write_draft(tmp_path, drop_fields={"Operating expenses"})
    items = build_review_items(draft)
    decisions = _decisions_for_items(draft, items)
    decisions_path = tmp_path / "decisions.json"
    decisions_path.write_text(decisions.model_dump_json(), encoding="utf-8")

    with pytest.raises(ValueError, match="No metric was extracted"):
        export_reviewed_run(tmp_path, decisions_path, tmp_path / "out.xlsx")

    artifacts = export_reviewed_run(
        tmp_path,
        decisions_path,
        tmp_path / "out.xlsx",
        allow_unreviewed=True,
    )
    assert artifacts.is_draft_unreviewed is True
    payload = json.loads(artifacts.json_path.read_text(encoding="utf-8"))
    assert any("Operating expenses" in warning for warning in payload["warnings"])


def test_export_cli_writes_artifact_paths(tmp_path: Path, capsys) -> None:
    run_dir = tmp_path / "run"
    draft_path = extract(INPUT_DIR, run_dir, mode="recorded")
    draft = DraftRun.model_validate_json(draft_path.read_text(encoding="utf-8"))
    decisions_path = run_dir / "human_review_decisions.json"
    decisions_path.write_text(
        build_acceptance_decisions(draft).model_dump_json(indent=2),
        encoding="utf-8",
    )

    exit_code = main(
        [
            "export",
            str(run_dir),
            "--decisions",
            str(decisions_path),
            "--out",
            str(tmp_path / "out.xlsx"),
        ]
    )
    captured = capsys.readouterr()

    assert exit_code == 0
    assert "out.xlsx" in captured.out
    assert "out.json" in captured.out
    assert "out.audit.md" in captured.out


def test_same_company_different_source_files_export_as_separate_rows(
    tmp_path: Path,
) -> None:
    source_a = "pdf_input/blackrock_q1_2025.pdf"
    source_b = "pdf_input/blackrock_q4_2025.pdf"
    metrics = _document_metrics(source_a, "BlackRock", "Q1 2025", revenue=5280)
    metrics.extend(_document_metrics(source_b, "BlackRock", "Q4 2025", revenue=7008))
    draft = DraftRun(
        run_id="same-company",
        created_at="2026-06-03T00:00:00Z",
        mode="live",
        model="test-model",
        reasoning_effort="low",
        documents=[
            {
                "source_file": source_a,
                "document_type": "earnings_report",
                "page_count": 10,
            },
            {
                "source_file": source_b,
                "document_type": "earnings_report",
                "page_count": 10,
            },
        ],
        classifications=[],
        selected_pages={source_a: [1], source_b: [1]},
        metrics=metrics,
    )
    (tmp_path / "draft_metrics.json").write_text(
        draft.model_dump_json(indent=2),
        encoding="utf-8",
    )
    items = build_review_items(draft)
    decisions = _decisions_for_items(draft, items)
    decisions_path = tmp_path / "decisions.json"
    decisions_path.write_text(decisions.model_dump_json(), encoding="utf-8")

    export_reviewed_run(tmp_path, decisions_path, tmp_path / "out.xlsx")

    worksheet = load_workbook(tmp_path / "out.xlsx").worksheets[0]
    rows = list(worksheet.iter_rows(min_row=2, max_row=3, values_only=True))
    assert rows[0][0:3] == ("BlackRock", "Q1 2025", "$5.3B")
    assert rows[1][0:3] == ("BlackRock", "Q4 2025", "$7B")


def test_batch_display_mode_uses_readable_units_and_missing_text(
    tmp_path: Path,
) -> None:
    source = "pdf_input/blackrock_q1_2025.pdf"
    metrics = _document_metrics(source, "BlackRock", "First Quarter 2025", 5276)
    for metric in metrics:
        if metric.metric_name == "Earnings per share":
            metric.value = 9.64
            metric.source_quote = "Diluted $ 9.64 $ 10.48"
        elif metric.metric_name == "Net income":
            metric.value = 1510.0
        elif metric.metric_name == "Operating income":
            metric.value = 1698.0
        elif metric.metric_name == "Gross margin":
            metric.value = None
            metric.unit = None
            metric.scale = None
            metric.needs_review = True
        elif metric.metric_name == "Operating expenses":
            metric.value = 3578.0
            metric.source_quote = "Total expense 3,578 3,035 543"
        elif metric.metric_name == "Buybacks and dividends":
            metric.value = (
                "$375 million share repurchases and $5.21 quarterly cash "
                "dividend per share"
            )
            metric.needs_review = True
    draft = DraftRun(
        run_id="batch-display",
        created_at="2026-06-03T00:00:00Z",
        mode="live",
        model="test-model",
        reasoning_effort="low",
        documents=[
            {
                "source_file": source,
                "document_type": "earnings_report",
                "page_count": 10,
            },
        ],
        classifications=[],
        selected_pages={source: [1]},
        metrics=metrics,
    )
    (tmp_path / "draft_metrics.json").write_text(
        draft.model_dump_json(indent=2),
        encoding="utf-8",
    )
    items = build_review_items(draft)
    decisions = _decisions_for_items(draft, items)
    decisions_path = tmp_path / "decisions.json"
    decisions_path.write_text(decisions.model_dump_json(), encoding="utf-8")

    export_reviewed_run(
        tmp_path,
        decisions_path,
        tmp_path / "out.xlsx",
        allow_unreviewed=True,
        display_mode="batch",
    )

    worksheet = load_workbook(tmp_path / "out.xlsx").worksheets[0]
    row = [cell.value for cell in worksheet[2]]
    assert row == [
        "BlackRock",
        "Q1 2025",
        "$5,276 million",
        "$9.64 diluted",
        "$1,510 million",
        "$1,698 million",
        "Not disclosed",
        "$3,578 million",
        "$375 million share repurchases; $5.21 per share cash dividend",
    ]


def _item(metric: MetricRow):
    draft = DraftRun(
        run_id="mapper",
        created_at="2026-05-31T00:00:00Z",
        mode="recorded",
        model="recorded",
        reasoning_effort=None,
        documents=[
            {
                "source_file": "test.pdf",
                "document_type": "earnings_report",
                "page_count": 1,
            }
        ],
        classifications=[],
        selected_pages={"test.pdf": [1]},
        metrics=[metric],
    )
    return build_review_items(draft)[0]


def _document_metrics(
    source_file: str,
    company: str,
    quarter: str,
    revenue: float,
) -> list[MetricRow]:
    values = {
        "Company Name": (company, None, None),
        "Quarter": (quarter, None, None),
        "Total revenue": (revenue, "USD", "millions"),
        "Earnings per share": (1.0, "USD/share", None),
        "Net income": (1000, "USD", "millions"),
        "Operating income": (1200, "USD", "millions"),
        "Gross margin": (10.0, "percentage points", None),
        "Operating expenses": (2000, "USD", "millions"),
        "Buybacks and dividends": ("$1B buybacks", None, None),
    }
    return [
        MetricRow(
            source_file=source_file,
            company=company,
            document_type="earnings_report",
            fiscal_period=quarter,
            metric_name=name,
            metric_category="template",
            value=value,
            unit=unit,
            scale=scale,
            source_page=1,
            source_quote=f"{name}: {value}",
            confidence=0.99,
            needs_review=False,
        )
        for name, (value, unit, scale) in values.items()
    ]


def _write_draft(tmp_path: Path, drop_fields: set[str] | None = None) -> DraftRun:
    drop_fields = drop_fields or set()
    all_metrics = [
        _metric("Company Name", "Tesla", unit=None, scale=None),
        _metric("Quarter", "Q2 2025", unit=None, scale=None),
        _metric("Total revenue", 22496),
        _metric("Earnings per share", 0.33, unit="USD/share", scale=None),
        _metric("Net income", 1172),
        _metric("Operating income", 923),
        _metric(
            "Gross margin",
            17.2,
            unit="percentage points",
            scale=None,
        ),
        _metric("Operating expenses", 2955),
        _metric(
            "Buybacks and dividends",
            None,
            unit=None,
            scale=None,
            needs_review=True,
        ),
    ]
    draft = DraftRun(
        run_id="gate-test",
        created_at="2026-05-31T00:00:00Z",
        mode="recorded",
        model="recorded",
        reasoning_effort=None,
        documents=[
            {
                "source_file": "test.pdf",
                "document_type": "earnings_report",
                "page_count": 1,
            }
        ],
        classifications=[],
        selected_pages={"test.pdf": [1]},
        metrics=[m for m in all_metrics if m.metric_name not in drop_fields],
    )
    (tmp_path / "draft_metrics.json").write_text(
        draft.model_dump_json(indent=2),
        encoding="utf-8",
    )
    return draft


def _metric(
    name: str,
    value: object,
    unit: str | None = "USD",
    scale: str | None = "millions",
    needs_review: bool = False,
) -> MetricRow:
    return MetricRow(
        company="Tesla",
        ticker="TSLA",
        document_type="earnings_report",
        fiscal_period="Q2 2025",
        metric_name=name,
        value=value,
        unit=unit,
        scale=scale,
        source_page=1,
        source_quote="source quote",
        confidence=0.5 if needs_review else 0.99,
        needs_review=needs_review,
        review_reason="Needs human review." if needs_review else None,
    )


def _decisions_for_items(
    draft: DraftRun,
    items: list,
) -> ReviewDecisionsFile:
    decisions = []
    for item in items:
        if item.requires_attention and item.value in (None, ""):
            status = "not_applicable"
            note = "Human resolved as not applicable."
        elif item.requires_attention:
            status = "approved"
            note = "Human approved flagged value."
        else:
            status = "approved"
            note = None
        decisions.append(
            ReviewDecision(
                metric_id=item.metric_id,
                review_status=status,
                reviewer_note=note,
                reviewed_at="2026-05-31T00:01:00Z",
                reviewer="Analyst",
            )
        )
    return ReviewDecisionsFile(
        run_id=draft.run_id,
        created_at="2026-05-31T00:01:00Z",
        is_demo=False,
        decisions=decisions,
    )
