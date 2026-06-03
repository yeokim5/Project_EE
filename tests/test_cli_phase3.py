import json
import time
from pathlib import Path

from openpyxl import load_workbook

from earnings_extractor.batch import run_batch
from earnings_extractor.cli import main
from earnings_extractor.config import OpenAIConfig
from earnings_extractor.extractor import LiveExtractionResult
from earnings_extractor.ingest import PageText
from earnings_extractor.pipeline import ProcessedDocument, find_pdf_inputs
from earnings_extractor.schema import (
    DocumentClassification,
    DraftRun,
    LLMUsage,
    MetricRow,
    MetricsBatch,
    PageClassification,
    SourceDocument,
)

ROOT = Path(__file__).resolve().parents[1]
TESLA = ROOT / "assesment_info" / "TSLA-Q2-2025-Update.pdf"
INPUT_DIR = ROOT / "assesment_info"


def test_find_pdf_inputs_accepts_single_pdf() -> None:
    assert find_pdf_inputs(TESLA) == [TESLA]


def test_find_pdf_inputs_accepts_directory() -> None:
    pdfs = find_pdf_inputs(INPUT_DIR)

    assert TESLA in pdfs
    assert all(path.suffix == ".pdf" for path in pdfs)


def test_extract_recorded_mode_succeeds_without_api_key(
    tmp_path: Path, capsys, monkeypatch
) -> None:
    def fail_config() -> None:
        raise AssertionError("recorded mode must not load OpenAI config")

    monkeypatch.delenv("OPENAI_API_KEY", raising=False)
    monkeypatch.setattr("earnings_extractor.pipeline.load_openai_config", fail_config)

    exit_code = main(
        [
            "extract",
            str(TESLA),
            "--out",
            str(tmp_path),
            "--mode",
            "recorded",
        ]
    )
    captured = capsys.readouterr()

    assert exit_code == 0
    assert "draft_metrics.json" in captured.out
    draft = DraftRun.model_validate_json(
        (tmp_path / "draft_metrics.json").read_text(encoding="utf-8")
    )
    assert draft.mode == "recorded"
    assert draft.llm_usage == []


def test_extract_live_mode_writes_openai_usage(
    tmp_path: Path, monkeypatch
) -> None:
    from earnings_extractor.pipeline import extract

    monkeypatch.setattr(
        "earnings_extractor.pipeline.load_openai_config",
        lambda: OpenAIConfig(
            api_key="test-key",
            model="test-model",
            reasoning_effort="low",
        ),
    )

    def fake_live_with_usage(*, pages, document_type, config, source_file):
        return LiveExtractionResult(
            metrics=MetricsBatch(metrics=[]),
            usage=LLMUsage(
                source_file=source_file,
                provider="openai",
                model=config.model,
                input_tokens=1000,
                output_tokens=200,
                total_tokens=1200,
                reasoning_tokens=50,
            ),
        )

    monkeypatch.setattr(
        "earnings_extractor.pipeline.extract_metrics_live_with_usage",
        fake_live_with_usage,
    )

    draft_path = extract(TESLA, tmp_path, mode="live")
    draft = DraftRun.model_validate_json(draft_path.read_text(encoding="utf-8"))

    assert draft.mode == "live"
    assert len(draft.llm_usage) == 1
    usage = draft.llm_usage[0]
    assert usage.source_file == str(TESLA)
    assert usage.input_tokens == 1000
    assert usage.output_tokens == 200
    assert usage.total_tokens == 1200
    assert usage.reasoning_tokens == 50


def test_uncertain_classification_still_extracts(monkeypatch, tmp_path: Path) -> None:
    from earnings_extractor.pipeline import process_single_pdf

    pdf_path = tmp_path / "ambiguous.pdf"
    pdf_path.write_bytes(b"%PDF-ambiguous")
    page = PageText(
        source_file=str(pdf_path),
        page_number=1,
        text="Q1 results Total revenue $10 Net income $2",
        char_count=40,
    )
    classification = DocumentClassification(
        source_file=str(pdf_path),
        document_type="unknown",
        page_count=1,
        pages=[
            PageClassification(
                page_number=1,
                style="mixed",
                char_count=40,
            )
        ],
    )
    seen: dict[str, str] = {}

    monkeypatch.setattr("earnings_extractor.pipeline.read_pdf_pages", lambda _: [page])
    monkeypatch.setattr("earnings_extractor.pipeline.read_pdf_metadata", lambda _: {})
    monkeypatch.setattr(
        "earnings_extractor.pipeline.classify_document",
        lambda _: classification,
    )
    monkeypatch.setattr(
        "earnings_extractor.pipeline.select_extraction_pages",
        lambda pages: pages,
    )

    def fake_extract_document_metrics(**kwargs):
        seen["document_type"] = kwargs["document_type"]
        return (
            MetricsBatch(
                metrics=[
                    MetricRow(
                        document_type="unknown",
                        metric_name="Total revenue",
                        metric_category="template",
                        value=10,
                        unit="USD",
                        scale="millions",
                        source_page=1,
                        source_quote="Total revenue $10",
                        confidence=0.9,
                        needs_review=False,
                    )
                ]
            ),
            None,
        )

    monkeypatch.setattr(
        "earnings_extractor.pipeline._extract_document_metrics",
        fake_extract_document_metrics,
    )

    processed = process_single_pdf(pdf_path, mode="live", config=None)

    assert seen["document_type"] == "earnings_report"
    assert processed.document is not None
    assert processed.document.document_type == "earnings_report"
    assert processed.metrics[0].document_type == "earnings_report"


def test_classification_type_does_not_change_extraction_hint(
    monkeypatch,
    tmp_path: Path,
) -> None:
    from earnings_extractor.pipeline import process_single_pdf

    pdf_path = tmp_path / "transcript.pdf"
    pdf_path.write_bytes(b"%PDF-transcript")
    page = PageText(
        source_file=str(pdf_path),
        page_number=1,
        text="Operator: Welcome to the earnings call. EPS was $1.00.",
        char_count=55,
    )
    classification = DocumentClassification(
        source_file=str(pdf_path),
        document_type="earnings_call_transcript",
        page_count=1,
        pages=[
            PageClassification(
                page_number=1,
                style="narrative",
                char_count=55,
            )
        ],
    )
    seen: dict[str, str] = {}

    monkeypatch.setattr("earnings_extractor.pipeline.read_pdf_pages", lambda _: [page])
    monkeypatch.setattr("earnings_extractor.pipeline.read_pdf_metadata", lambda _: {})
    monkeypatch.setattr(
        "earnings_extractor.pipeline.classify_document",
        lambda _: classification,
    )
    monkeypatch.setattr(
        "earnings_extractor.pipeline.select_extraction_pages",
        lambda pages: pages,
    )

    def fake_extract_document_metrics(**kwargs):
        seen["document_type"] = kwargs["document_type"]
        return (
            MetricsBatch(
                metrics=[
                    MetricRow(
                        document_type="earnings_call_transcript",
                        metric_name="Earnings per share",
                        metric_category="template",
                        value=1.0,
                        unit="USD/share",
                        scale="ones",
                        source_page=1,
                        source_quote="EPS was $1.00.",
                        confidence=0.9,
                        needs_review=False,
                    )
                ]
            ),
            None,
        )

    monkeypatch.setattr(
        "earnings_extractor.pipeline._extract_document_metrics",
        fake_extract_document_metrics,
    )

    processed = process_single_pdf(pdf_path, mode="live", config=None)

    assert seen["document_type"] == "earnings_report"
    assert processed.document is not None
    assert processed.document.document_type == "earnings_report"


def test_heartbeat_reports_long_running_operation() -> None:
    from earnings_extractor.pipeline import _with_heartbeat

    messages: list[str] = []

    def slow_operation():
        time.sleep(0.03)
        return MetricsBatch(metrics=[]), None

    result, usage = _with_heartbeat(
        slow_operation,
        progress=messages.append,
        message="running",
        enabled=True,
        interval_seconds=0.01,
    )

    assert result.metrics == []
    assert usage is None
    assert any(message.startswith("running") for message in messages)


def test_metrics_keep_exact_source_file_for_export_grouping(
    monkeypatch,
    tmp_path: Path,
) -> None:
    from earnings_extractor.pipeline import process_single_pdf

    pdf_path = tmp_path / "blackrock_q1_2025.pdf"
    pdf_path.write_bytes(b"%PDF")
    page = PageText(
        source_file=str(pdf_path),
        page_number=1,
        text="Q1 2025 results Total revenue $10",
        char_count=34,
    )

    monkeypatch.setattr("earnings_extractor.pipeline.read_pdf_pages", lambda _: [page])
    monkeypatch.setattr("earnings_extractor.pipeline.read_pdf_metadata", lambda _: {})
    monkeypatch.setattr(
        "earnings_extractor.pipeline.select_extraction_pages",
        lambda pages: pages,
    )
    monkeypatch.setattr(
        "earnings_extractor.pipeline._extract_document_metrics",
        lambda **_: (
            MetricsBatch(
                metrics=[
                    MetricRow(
                        document_type="earnings_report",
                        metric_name="Total revenue",
                        metric_category="template",
                        value=10,
                        unit="USD",
                        scale="millions",
                        source_page=1,
                        source_quote="Total revenue $10",
                        confidence=0.9,
                        needs_review=False,
                    )
                ]
            ),
            None,
        ),
    )

    processed = process_single_pdf(pdf_path, mode="live", config=None)

    assert processed.metrics[0].source_file == str(pdf_path)


def test_batch_cli_outputs_client_first_review_tabs_and_short_summary(
    monkeypatch,
    tmp_path: Path,
) -> None:
    input_dir = tmp_path / "pdfs"
    input_dir.mkdir()
    pdf_path = input_dir / "sample.pdf"
    pdf_path.write_bytes(b"%PDF")

    def fake_process_single_pdf(pdf_path, mode, config, progress=None):
        metrics = _batch_metrics(str(pdf_path))
        return ProcessedDocument(
            source_file=str(pdf_path),
            document=SourceDocument(
                source_file=str(pdf_path),
                document_type="earnings_report",
                page_count=1,
            ),
            classification=DocumentClassification(
                source_file=str(pdf_path),
                document_type="earnings_report",
                page_count=1,
                pages=[],
            ),
            selected_pages=[1],
            usage=None,
            metrics=metrics,
        )

    monkeypatch.setattr(
        "earnings_extractor.batch.process_single_pdf",
        fake_process_single_pdf,
    )

    out_xlsx = tmp_path / "out.xlsx"
    summary = run_batch(input_dir, out_xlsx, mode="recorded")

    assert "Review summary:" not in summary.as_text()
    assert "Needs attention:" not in summary.as_text()
    assert "Review details are in the workbook" in summary.as_text()

    workbook = load_workbook(out_xlsx)
    assert workbook.sheetnames[:4] == [
        "Review Instructions",
        "Extraction Draft",
        "Review Queue",
        "Batch Status",
    ]
    instructions = workbook["Review Instructions"]
    assert instructions["A1"].value == "Workbook state"
    assert (
        instructions["B4"].value
        == "OK: populated value passed deterministic checks."
    )
    queue = workbook["Review Queue"]
    headers = [cell.value for cell in queue[1]]
    assert headers == [
        "Status",
        "Source file",
        "Company",
        "Quarter",
        "Field",
        "Displayed value",
        "Reason",
        "Source page",
        "Source quote",
    ]
    rows = list(queue.iter_rows(min_row=2, values_only=True))
    assert any(row[0] == "OK" and row[4] == "Total revenue" for row in rows)
    assert any(
        row[0] == "NEEDS REVIEW" and row[4] == "Buybacks and dividends"
        for row in rows
    )
    assert any(row[0] == "NOT DISCLOSED" and row[4] == "Gross margin" for row in rows)

    assert "Start on Extraction Draft" in instructions["B2"].value

    client = workbook["Extraction Draft"]
    assert client["C2"].fill.fgColor.rgb == "00E2F0D9"
    assert client["G2"].comment is not None
    assert "NOT DISCLOSED" in client["G2"].comment.text
    assert client["G2"].fill.fgColor.rgb == "00E7E6E6"
    assert client["I2"].comment is not None
    assert "NEEDS REVIEW" in client["I2"].comment.text
    assert client["I2"].fill.fgColor.rgb == "00FFF2CC"


def test_inspect_summarizes_draft_file(tmp_path: Path, capsys) -> None:
    draft_path = _write_sample_draft(tmp_path)

    exit_code = main(["inspect", str(draft_path)])
    captured = capsys.readouterr()

    assert exit_code == 0
    assert "Metrics: 1" in captured.out
    assert "With evidence: 1/1" in captured.out


def test_eval_prints_field_level_score(tmp_path: Path, capsys) -> None:
    draft_path = _write_sample_draft(tmp_path)

    exit_code = main(
        [
            "eval",
            "--draft",
            str(draft_path),
            "--document-id",
            "tesla_q2_2025",
        ]
    )
    captured = capsys.readouterr()

    assert exit_code == 0
    assert "Document: tesla_q2_2025" in captured.out
    assert "Score:" in captured.out


def test_eval_without_min_accuracy_is_report_only(tmp_path: Path, capsys) -> None:
    draft_path = _write_low_score_draft(tmp_path)

    exit_code = main(
        [
            "eval",
            "--draft",
            str(draft_path),
            "--document-id",
            "tesla_q2_2025",
        ]
    )
    captured = capsys.readouterr()

    assert exit_code == 0
    assert "Score:" in captured.out
    assert "below minimum" not in captured.err


def test_eval_min_accuracy_passes_for_recorded_tesla(tmp_path: Path) -> None:
    from earnings_extractor.pipeline import extract

    draft_path = extract(TESLA, tmp_path, mode="recorded")

    exit_code = main(
        [
            "eval",
            "--draft",
            str(draft_path),
            "--document-id",
            "tesla_q2_2025",
            "--min-accuracy",
            "0.9",
        ]
    )

    assert exit_code == 0


def test_eval_min_accuracy_fails_below_threshold(tmp_path: Path, capsys) -> None:
    draft_path = _write_low_score_draft(tmp_path)

    exit_code = main(
        [
            "eval",
            "--draft",
            str(draft_path),
            "--document-id",
            "tesla_q2_2025",
            "--min-accuracy",
            "0.9",
        ]
    )
    captured = capsys.readouterr()

    assert exit_code == 1
    assert "Accuracy" in captured.err
    assert "below minimum" in captured.err


def test_eval_min_accuracy_rejects_invalid_threshold(tmp_path: Path, capsys) -> None:
    draft_path = _write_sample_draft(tmp_path)

    exit_code = main(
        [
            "eval",
            "--draft",
            str(draft_path),
            "--document-id",
            "tesla_q2_2025",
            "--min-accuracy",
            "1.1",
        ]
    )
    captured = capsys.readouterr()

    assert exit_code == 2
    assert "between 0.0 and 1.0" in captured.err


def _write_sample_draft(tmp_path: Path) -> Path:
    draft = DraftRun(
        run_id="test",
        created_at="2026-05-29T00:00:00Z",
        mode="live",
        model="test-model",
        reasoning_effort="low",
        documents=[
            {
                "source_file": str(TESLA),
                "document_type": "earnings_report",
                "page_count": 30,
            }
        ],
        classifications=[],
        selected_pages={str(TESLA): [4]},
        metrics=[
            {
                "company": "Tesla",
                "ticker": "TSLA",
                "document_type": "earnings_report",
                "fiscal_period": "Q2 2025",
                "report_date": None,
                "metric_name": "Total revenue",
                "metric_category": "income_statement",
                "segment": None,
                "value": 22496,
                "unit": "USD",
                "scale": "millions",
                "period": "quarter",
                "gaap_or_non_gaap": "GAAP",
                "year_over_year_change": None,
                "source_page": 4,
                "source_quote": (
                    "Total revenues 25,500 25,182 25,707 19,335 22,496 -12%"
                ),
                "confidence": 0.9,
                "needs_review": False,
                "review_reason": None,
                "review_status": "pending",
                "reviewer_note": None,
            }
        ],
    )
    draft_path = tmp_path / "draft_metrics.json"
    draft_path.write_text(
        json.dumps(draft.model_dump(mode="json"), indent=2),
        encoding="utf-8",
    )
    return draft_path


def _batch_metrics(source_file: str) -> list[MetricRow]:
    base = {
        "Company Name": ("ExampleCo", None, None, False, None),
        "Quarter": ("Q1 2026", None, None, False, None),
        "Total revenue": (1000, "USD", "millions", False, None),
        "Earnings per share": (1.23, "USD/share", None, False, None),
        "Net income": (100, "USD", "millions", False, None),
        "Operating income": (120, "USD", "millions", False, None),
        "Gross margin": (
            None,
            None,
            None,
            True,
            "Gross margin was not disclosed.",
        ),
        "Operating expenses": (880, "USD", "millions", False, None),
        "Buybacks and dividends": (
            "$10 million share repurchases",
            None,
            None,
            True,
            "Combined capital return field needs review.",
        ),
    }
    return [
        MetricRow(
            source_file=source_file,
            company="ExampleCo",
            document_type="earnings_report",
            fiscal_period="Q1 2026",
            metric_name=name,
            metric_category="template",
            value=value,
            unit=unit,
            scale=scale,
            source_page=1,
            source_quote=f"{name}: {value}",
            confidence=0.99,
            needs_review=needs_review,
            review_reason=reason,
        )
        for name, (value, unit, scale, needs_review, reason) in base.items()
    ]


def _write_low_score_draft(tmp_path: Path) -> Path:
    draft_path = _write_sample_draft(tmp_path)
    draft = DraftRun.model_validate_json(draft_path.read_text(encoding="utf-8"))
    draft.metrics[0].value = 1.0
    draft_path.write_text(
        json.dumps(draft.model_dump(mode="json"), indent=2),
        encoding="utf-8",
    )
    return draft_path
