"""Unit + integration tests for the Phase 11 web orchestration layer."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from earnings_extractor.export import export_reviewed_run
from earnings_extractor.pipeline import extract
from earnings_extractor.review import build_review_items
from earnings_extractor.schema import (
    DocumentClassification,
    DraftRun,
    MetricRow,
    SourceDocument,
    new_run_id,
    utc_now_iso,
)
from scripts.web_lib import (
    build_decisions_file,
    evidence_for,
    merge_drafts,
    template_metric_payloads,
)

ROOT = Path(__file__).resolve().parents[1]
INPUT_DIR = ROOT / "assesment_info"
TESLA = INPUT_DIR / "TSLA-Q2-2025-Update.pdf"
CITI = INPUT_DIR / "citi_earnings_q12025.pdf"


def _draft(source_file: str, metric_names: list[str]) -> DraftRun:
    metrics = [
        MetricRow(
            metric_name=name,
            value=float(index + 1),
            unit="USD",
            scale="millions",
            source_page=1,
            source_quote=f"{name} was {index + 1}.",
            confidence=0.9,
            needs_review=False,
        )
        for index, name in enumerate(metric_names)
    ]
    return DraftRun(
        run_id=new_run_id(),
        created_at=utc_now_iso(),
        mode="recorded",
        documents=[
            SourceDocument(
                source_file=source_file,
                document_type="earnings_report",
                page_count=1,
            )
        ],
        classifications=[
            DocumentClassification(
                source_file=source_file,
                document_type="earnings_report",
                page_count=1,
                pages=[],
            )
        ],
        selected_pages={source_file: [1]},
        metrics=metrics,
    )


def test_merge_drafts_offsets_and_fresh_run_id() -> None:
    a = _draft("a.pdf", ["Total revenue", "Net income"])
    b = _draft("b.pdf", ["Total revenue"])

    merged, offsets = merge_drafts([a, b])

    assert offsets == [0, 2]
    assert len(merged.metrics) == 3
    assert merged.run_id not in {a.run_id, b.run_id}
    assert [doc.source_file for doc in merged.documents] == ["a.pdf", "b.pdf"]


def test_build_decisions_maps_by_document_and_local_index() -> None:
    a = _draft("a.pdf", ["Total revenue", "Net income"])
    b = _draft("b.pdf", ["Total revenue", "Net income"])
    merged, offsets = merge_drafts([a, b])

    # Reject document b's "Net income" (local index 1 in document 1).
    per_doc = [
        [],
        [{"metric_index": 1, "review_status": "rejected", "note": "wrong line"}],
    ]
    decisions = build_decisions_file(merged, offsets, per_doc, reviewer="Tester")

    items = build_review_items(merged)
    by_id = {d.metric_id: d for d in decisions.decisions}
    # global index 3 == document b, local 1 == its "Net income"
    target = next(item for item in items if item.metric_index == 3)
    assert by_id[target.metric_id].review_status == "rejected"
    assert by_id[target.metric_id].reviewer_note == "wrong line"
    # everything else defaulted to approved
    other = next(item for item in items if item.metric_index == 0)
    assert by_id[other.metric_id].review_status == "approved"
    assert decisions.is_demo is False


def test_evidence_for_returns_fractions_on_golden_doc() -> None:
    # A phrase that appears verbatim in the Citi transcript cassette quotes.
    location = evidence_for(CITI, 1, "Citi")
    assert "rects" in location and "matched" in location
    for rect in location["rects"]:
        assert 0.0 <= rect["left"] <= 1.0
        assert 0.0 <= rect["top"] <= 1.0


def test_full_merge_export_with_edited_value(tmp_path: Path) -> None:
    """Two single-doc recorded drafts merge into one gated final workbook."""

    tesla_dir = tmp_path / "tesla"
    citi_dir = tmp_path / "citi"
    tesla_draft = DraftRun.model_validate_json(
        extract(TESLA, tesla_dir, mode="recorded").read_text(encoding="utf-8")
    )
    citi_draft = DraftRun.model_validate_json(
        extract(CITI, citi_dir, mode="recorded").read_text(encoding="utf-8")
    )

    # The client attaches evidence + metric_index payloads for review.
    payloads = template_metric_payloads(tesla_draft, TESLA)
    assert payloads and all("evidence" in p for p in payloads)

    # Simulate a human editing Tesla's Total revenue to a new value.
    revenue_index = next(
        index
        for index, metric in enumerate(tesla_draft.metrics)
        if metric.metric_name == "Total revenue"
    )
    tesla_draft.metrics[revenue_index].value = 99999.0
    tesla_draft.metrics[revenue_index].unit = "USD"
    tesla_draft.metrics[revenue_index].scale = "millions"

    merged, offsets = merge_drafts([tesla_draft, citi_draft])
    decisions = build_decisions_file(merged, offsets, [[], []], reviewer="Tester")

    run_dir = tmp_path / "run"
    run_dir.mkdir()
    (run_dir / "draft_metrics.json").write_text(
        merged.model_dump_json(indent=2), encoding="utf-8"
    )
    decisions_path = run_dir / "review_decisions.json"
    decisions_path.write_text(decisions.model_dump_json(indent=2), encoding="utf-8")

    artifacts = export_reviewed_run(
        run_dir=run_dir,
        decisions_path=decisions_path,
        out_path=tmp_path / "merged.xlsx",
        allow_unreviewed=False,
    )
    assert artifacts.is_draft_unreviewed is False

    workbook = load_workbook(artifacts.xlsx_path)
    sheet = workbook.worksheets[0]
    assert sheet.title == "Client Template"
    # Two documents -> two client rows.
    assert sheet.max_row >= 3
    # The edited Tesla revenue (99999 -> $100B) flows into the workbook.
    revenue_cells = [
        sheet.cell(row=row, column=3).value for row in range(2, sheet.max_row + 1)
    ]
    assert "$100B" in revenue_cells
