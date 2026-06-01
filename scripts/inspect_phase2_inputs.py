#!/usr/bin/env python3
"""Verify Phase 2 input readability and template contract."""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from earnings_extractor.ingest import PageText, read_pdf_pages  # noqa: E402
from evaluation.golden_metrics import (  # noqa: E402
    AUXILIARY_METRICS,
    PRIMARY_FIELDS,
    TEMPLATE_FIELDS,
)
from evaluation.tolerances import normalize_whitespace  # noqa: E402

TEMPLATE_PATH = ROOT / "assesment_info" / "EarningsSample (1).xlsx"

EXPECTED_SAMPLE_ROW = {
    "Company Name": ("Amazon LLC", "General"),
    "Quarter": ("Q1 2025", "General"),
    "Total revenue": ("$150B", "General"),
    "Earnings per share": (0.8, '"$"#,##0.00'),
    "Net income": ("$8.3B", "General"),
    "Operating income": ("$12.5B", "General"),
    "Gross margin": (0.46, "0%"),
    "Operating expenses": ("$142.5B", "General"),
    "Buybacks and dividends": ("$0.5B Buybacks, $0 Dividends", "General"),
}


def quote_present(page_text: str, quote: str) -> bool:
    return normalize_whitespace(quote) in normalize_whitespace(page_text)


def pages_by_number(
    pdf_path: Path, page_numbers: set[int] | None = None
) -> dict[int, PageText]:
    return {
        page.page_number: page for page in read_pdf_pages(pdf_path, page_numbers)
    }


def check_pdf_evidence() -> list[str]:
    failures: list[str] = []
    page_cache: dict[Path, dict[int, PageText]] = {}
    evidence_rows = [
        row for row in PRIMARY_FIELDS if row.status == "expected_value"
    ] + list(AUXILIARY_METRICS)
    cited_pages_by_pdf: dict[Path, set[int]] = {}
    for row in evidence_rows:
        if row.source_page is not None:
            cited_pages_by_pdf.setdefault(row.source_file, set()).add(row.source_page)

    for row in evidence_rows:
        source_page = row.source_page
        source_quote = row.source_quote
        if source_page is None or not source_quote:
            failures.append(f"{row.document_id}:{row.field_name} is missing evidence")
            continue

        pdf_path = row.source_file
        page_cache.setdefault(
            pdf_path, pages_by_number(pdf_path, cited_pages_by_pdf.get(pdf_path))
        )
        page = page_cache[pdf_path].get(source_page)
        label = getattr(row, "field_name", getattr(row, "metric_name", "unknown"))
        if page is None:
            failures.append(f"{row.document_id}:{label} page {source_page} not found")
            continue
        if not quote_present(page.text, source_quote):
            failures.append(
                f"{row.document_id}:{label} quote not found on page {source_page}"
            )

    return failures


def check_excel_template() -> list[str]:
    failures: list[str] = []
    workbook = load_workbook(TEMPLATE_PATH, data_only=False)
    sheet = workbook.active

    headers = [sheet.cell(row=1, column=index).value for index in range(1, 10)]
    if tuple(headers) != TEMPLATE_FIELDS:
        failures.append(f"Template headers differ: {headers!r}")

    for index, header in enumerate(TEMPLATE_FIELDS, start=1):
        expected_value, expected_format = EXPECTED_SAMPLE_ROW[header]
        cell = sheet.cell(row=2, column=index)
        if cell.value != expected_value:
            failures.append(
                f"Sample row {header!r} value {cell.value!r} != {expected_value!r}"
            )
        if cell.number_format != expected_format:
            failures.append(
                f"Sample row {header!r} format "
                f"{cell.number_format!r} != {expected_format!r}"
            )

    return failures


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Verify Phase 2 PDF readability and Excel template format."
    )
    parser.add_argument(
        "--check",
        action="store_true",
        help="Exit nonzero if any Phase 2 input check fails.",
    )
    return parser.parse_args()


def main() -> int:
    parse_args()
    failures = check_pdf_evidence() + check_excel_template()

    if failures:
        for failure in failures:
            print(f"FAIL: {failure}")
        return 1

    print("Phase 2 inputs are readable and template format checks passed.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
