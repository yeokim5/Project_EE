"""One-command batch lane: a folder of PDFs in, one Excel workbook out.

This is the path for "I have 20 PDFs and I just want a spreadsheet." It reuses
the exact same per-document engine as the review-first ``extract``/``export``
workflow -- ingest, classify, LLM draft, deterministic normalize/validate, and
the gated export -- but wraps each PDF so that one bad file (corrupt, encrypted,
scanned-image, or simply not an earnings document) is recorded and skipped
instead of sinking the whole batch.

Every value that *was* extracted is written to the client sheet immediately and
the workbook is clearly stamped ``DRAFT/UNREVIEWED`` (amber tab), because nothing
here has been through human review yet. A "Batch Status" sheet lists every input
file and what happened to it, so a missing company is visible rather than silent.
"""

from __future__ import annotations

from collections.abc import Callable
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill

from earnings_extractor.config import load_openai_config
from earnings_extractor.export import (
    export_reviewed_run,
    map_metric_to_client_cell,
)
from earnings_extractor.pipeline import (
    build_draft_run,
    find_pdf_inputs,
    process_single_pdf,
)
from earnings_extractor.review import build_review_items
from earnings_extractor.schema import (
    TEMPLATE_FIELDS,
    DraftRun,
    LLMUsage,
    SourceDocument,
)
from scripts.make_acceptance_decisions import build_acceptance_decisions

STATUS_EXTRACTED = "extracted"
STATUS_SKIPPED_NOT_EARNINGS = "skipped (not an earnings doc)"
STATUS_FAILED = "failed"
REVIEW_OK = "OK"
REVIEW_NEEDS_REVIEW = "NEEDS REVIEW"
REVIEW_NOT_DISCLOSED = "NOT DISCLOSED"

_CLIENT_SHEET_PREFIX = "Client Template"
_EXTRACTION_DRAFT_SHEET = "Extraction Draft"
_COMMENT_AUTHOR = "earnings_extractor"
_NEEDS_REVIEW_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFF2CC",
)
_NOT_DISCLOSED_FILL = PatternFill(
    fill_type="solid",
    fgColor="E7E6E6",
)
_OK_FILL = PatternFill(
    fill_type="solid",
    fgColor="E2F0D9",
)


@dataclass
class FileStatus:
    file: str
    status: str
    document_type: str = ""
    metrics_found: int = 0
    needs_review: int = 0
    error: str = ""


@dataclass(frozen=True)
class ReviewQueueRow:
    status: str
    source_file: str
    company: str
    quarter: str
    field: str
    displayed_value: str
    reason: str
    source_page: int
    source_quote: str


@dataclass
class BatchSummary:
    out_xlsx: Path | None
    total: int = 0
    extracted: int = 0
    skipped: int = 0
    failed: int = 0
    statuses: list[FileStatus] = field(default_factory=list)
    review_queue: list[ReviewQueueRow] = field(default_factory=list)

    def as_text(self) -> str:
        lines = [
            f"Processed {self.total} PDF(s): "
            f"{self.extracted} extracted, {self.skipped} skipped, "
            f"{self.failed} failed.",
        ]
        for status in self.statuses:
            detail = status.document_type or status.error or ""
            suffix = f" -- {detail}" if detail else ""
            lines.append(f"  [{status.status}] {status.file}{suffix}")
        if self.out_xlsx is not None:
            lines.append(f"Workbook: {self.out_xlsx}")
            lines.append(
                "Review details are in the workbook: see Review Instructions "
                "and Review Queue."
            )
        return "\n".join(lines) + "\n"


def run_batch(
    input_path: Path,
    out_xlsx: Path,
    mode: str = "live",
    progress: Callable[[str], None] | None = None,
) -> BatchSummary:
    """Extract every PDF under ``input_path`` into a single workbook.

    Per-PDF failures are caught and reported; they never abort the batch.
    """

    if mode not in {"live", "recorded"}:
        raise ValueError(f"Unsupported mode: {mode}")

    if not input_path.exists():
        raise FileNotFoundError(
            f"Input folder not found: {input_path}\n"
            "Create it and drop your PDFs inside, e.g.:\n"
            f"    mkdir -p {input_path} && cp *.pdf {input_path}/"
        )

    pdf_paths = find_pdf_inputs(input_path)
    if not pdf_paths:
        raise ValueError(f"No PDF files found in {input_path}")

    config = load_openai_config() if mode == "live" else None
    _emit(progress, f"Found {len(pdf_paths)} PDF(s) in {input_path}.")

    documents: list[SourceDocument] = []
    classifications: list = []
    selected_pages: dict[str, list[int]] = {}
    llm_usage: list[LLMUsage] = []
    all_metrics: list = []
    statuses: list[FileStatus] = []

    for index, pdf_path in enumerate(pdf_paths, start=1):
        name = pdf_path.name
        prefix = f"[{index}/{len(pdf_paths)}] {name}"
        _emit(progress, f"{prefix}: started.")
        try:
            processed = process_single_pdf(
                pdf_path,
                mode,
                config,
                progress=lambda message, prefix=prefix: _emit(
                    progress,
                    f"{prefix}: {message}",
                ),
            )
        except Exception as exc:  # noqa: BLE001 -- isolate one bad PDF
            statuses.append(
                FileStatus(file=name, status=STATUS_FAILED, error=str(exc))
            )
            _emit(progress, f"{prefix}: failed -- {exc}")
            continue

        if processed.document is None:
            statuses.append(
                FileStatus(file=name, status=STATUS_SKIPPED_NOT_EARNINGS)
            )
            _emit(progress, f"{prefix}: skipped (not an earnings doc)")
            continue

        if processed.usage is not None:
            llm_usage.append(processed.usage)
        documents.append(processed.document)
        classifications.append(processed.classification)
        selected_pages[str(pdf_path)] = processed.selected_pages
        all_metrics.extend(processed.metrics)
        statuses.append(
            FileStatus(
                file=name,
                status=STATUS_EXTRACTED,
                document_type=processed.document.document_type,
                metrics_found=len(processed.metrics),
                needs_review=sum(
                    1 for metric in processed.metrics if metric.needs_review
                ),
            )
        )
        _emit(
            progress,
            (
                f"{prefix}: done - {len(processed.metrics)} metric(s), "
                f"{sum(1 for metric in processed.metrics if metric.needs_review)} "
                "need review."
            ),
        )

    summary = BatchSummary(
        out_xlsx=None,
        total=len(pdf_paths),
        extracted=sum(1 for s in statuses if s.status == STATUS_EXTRACTED),
        skipped=sum(
            1 for s in statuses if s.status == STATUS_SKIPPED_NOT_EARNINGS
        ),
        failed=sum(1 for s in statuses if s.status == STATUS_FAILED),
        statuses=statuses,
    )

    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    _emit(progress, f"Writing workbook to {out_xlsx}...")

    if not documents:
        # Nothing usable was extracted. Still produce a workbook so the user
        # gets a single, openable artifact that explains what happened.
        _write_status_only_workbook(out_xlsx, statuses)
        summary.out_xlsx = out_xlsx
        _emit(progress, f"Wrote {out_xlsx}.")
        return summary

    run_dir = out_xlsx.parent
    draft = build_draft_run(
        mode=mode,
        config=config,
        documents=documents,
        classifications=classifications,
        selected_pages=selected_pages,
        llm_usage=llm_usage,
        metrics=all_metrics,
    )
    draft_path = run_dir / "draft_metrics.json"
    draft_path.write_text(
        _dump_draft(draft),
        encoding="utf-8",
    )

    decisions = build_acceptance_decisions(draft)
    decisions_path = run_dir / "batch_decisions.json"
    decisions_path.write_text(
        _dump_model(decisions),
        encoding="utf-8",
    )
    review_queue = _build_review_queue(draft)
    summary.review_queue = review_queue

    # allow_unreviewed=True: populate every extracted value now and stamp the
    # workbook DRAFT/UNREVIEWED. A company missing a metric leaves a blank cell
    # (recorded in warnings + the status sheet) -- it never blocks the export.
    export_reviewed_run(
        run_dir,
        decisions_path=decisions_path,
        out_path=out_xlsx,
        allow_unreviewed=True,
        display_mode="batch",
    )

    _append_batch_review_artifacts(out_xlsx, statuses, review_queue)
    summary.out_xlsx = out_xlsx
    _emit(progress, f"Wrote {out_xlsx}.")
    return summary


def _emit(progress: Callable[[str], None] | None, message: str) -> None:
    if progress is not None:
        progress(message)


def _dump_draft(draft) -> str:
    import json

    return json.dumps(draft.model_dump(mode="json"), indent=2) + "\n"


def _dump_model(model) -> str:
    import json

    return json.dumps(model.model_dump(mode="json"), indent=2) + "\n"


def _build_review_queue(draft: DraftRun) -> list[ReviewQueueRow]:
    rows: list[ReviewQueueRow] = []
    items = [
        item
        for item in build_review_items(draft)
        if item.metric_name in TEMPLATE_FIELDS
    ]
    for item in items:
        status = _review_status(item)
        rows.append(
            ReviewQueueRow(
                status=status,
                source_file=item.source_file,
                company=item.company or "",
                quarter=str(_item_value_by_name(items, item.source_file, "Quarter")),
                field=item.metric_name,
                displayed_value=_display_value(item),
                reason=_review_reason(item, status),
                source_page=item.source_page,
                source_quote=item.source_quote,
            )
        )
    return rows


def _item_value_by_name(items: list[Any], source_file: str, metric_name: str) -> str:
    for item in items:
        if item.source_file == source_file and item.metric_name == metric_name:
            return "" if item.value in (None, "") else str(item.value)
    return ""


def _review_status(item: Any) -> str:
    if item.value in (None, ""):
        return REVIEW_NOT_DISCLOSED
    if item.needs_review:
        return REVIEW_NEEDS_REVIEW
    return REVIEW_OK


def _review_reason(item: Any, status: str) -> str:
    if item.review_reason:
        return item.review_reason
    if status == REVIEW_NOT_DISCLOSED:
        return "Field was not disclosed in selected source pages."
    return ""


def _display_value(item: Any) -> str:
    if item.value in (None, ""):
        return ""
    try:
        value = map_metric_to_client_cell(item, display_mode="batch")
    except ValueError:
        value = item.value
    return "" if value is None else str(value)


_STATUS_HEADERS = [
    "Source file",
    "Status",
    "Document type",
    "Metrics found",
    "Needs review",
    "Error",
]


def _status_row(status: FileStatus) -> list:
    return [
        status.file,
        status.status,
        status.document_type,
        status.metrics_found if status.status == STATUS_EXTRACTED else "",
        status.needs_review if status.status == STATUS_EXTRACTED else "",
        status.error,
    ]


def _append_batch_review_artifacts(
    out_xlsx: Path,
    statuses: list[FileStatus],
    review_queue: list[ReviewQueueRow],
) -> None:
    workbook = load_workbook(out_xlsx)
    _rename_client_sheet(workbook)
    _append_review_instructions_sheet(workbook)
    _append_review_queue_sheet(workbook, review_queue)
    _append_status_sheet(workbook, statuses)
    _annotate_client_sheet(workbook, review_queue)
    _order_batch_sheets(workbook)
    workbook.save(out_xlsx)


def _rename_client_sheet(workbook: Any) -> None:
    worksheet = _client_sheet(workbook)
    if worksheet is None:
        return
    worksheet.title = _EXTRACTION_DRAFT_SHEET


def _order_batch_sheets(workbook: Any) -> None:
    priority = [
        "Review Instructions",
        _EXTRACTION_DRAFT_SHEET,
        "Review Queue",
        "Batch Status",
    ]
    ordered = []
    for title in priority:
        if title in workbook.sheetnames:
            ordered.append(workbook[title])
    ordered_ids = {id(sheet) for sheet in ordered}
    ordered.extend(
        sheet for sheet in workbook.worksheets if id(sheet) not in ordered_ids
    )
    workbook._sheets = ordered


def _append_status_sheet(workbook: Any, statuses: list[FileStatus]) -> None:
    if "Batch Status" in workbook.sheetnames:
        del workbook["Batch Status"]
    worksheet = workbook.create_sheet("Batch Status")
    worksheet.append(_STATUS_HEADERS)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
    for status in statuses:
        worksheet.append(_status_row(status))


def _append_review_instructions_sheet(workbook: Any) -> None:
    if "Review Instructions" in workbook.sheetnames:
        del workbook["Review Instructions"]
    worksheet = workbook.create_sheet("Review Instructions")
    rows = [
        ["Workbook state", "DRAFT/UNREVIEWED"],
        [
            "How to use",
            (
                "Start on Extraction Draft. Colored cells have comments. "
                "Use Review Queue for every field's status, reason, source page, "
                "and source quote."
            ),
        ],
        ["Color", "Meaning"],
        ["Green", "OK: populated value passed deterministic checks."],
        [
            "Yellow",
            (
                "NEEDS REVIEW: populated value has a review flag, such as "
                "low confidence, citation mismatch, or auto-enriched text."
            ),
        ],
        [
            "Gray",
            (
                "NOT DISCLOSED: field is blank because it was not found or not "
                "applicable in the selected source pages."
            ),
        ],
    ]
    for row in rows:
        worksheet.append(row)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
    worksheet["A4"].fill = _OK_FILL
    worksheet["A5"].fill = _NEEDS_REVIEW_FILL
    worksheet["A6"].fill = _NOT_DISCLOSED_FILL


_REVIEW_QUEUE_HEADERS = [
    "Status",
    "Source file",
    "Company",
    "Quarter",
    "Field",
    "Displayed value",
    "Reason",
    "Source page",
    "Source quote",
]


def _append_review_queue_sheet(
    workbook: Any,
    review_queue: list[ReviewQueueRow],
) -> None:
    if "Review Queue" in workbook.sheetnames:
        del workbook["Review Queue"]
    worksheet = workbook.create_sheet("Review Queue")
    worksheet.append(_REVIEW_QUEUE_HEADERS)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
    for row in review_queue:
        worksheet.append(
            [
                row.status,
                row.source_file,
                row.company,
                row.quarter,
                row.field,
                row.displayed_value,
                row.reason,
                row.source_page,
                row.source_quote,
            ]
        )


def _annotate_client_sheet(
    workbook: Any,
    review_queue: list[ReviewQueueRow],
) -> None:
    worksheet = _client_sheet(workbook)
    if worksheet is None:
        return
    rows_by_source = _review_rows_by_source(review_queue)
    for row_number, source_file in enumerate(rows_by_source, start=2):
        field_rows = {row.field: row for row in rows_by_source[source_file]}
        for column_number, metric_name in enumerate(TEMPLATE_FIELDS, start=1):
            review_row = field_rows.get(metric_name)
            if review_row is None:
                continue
            cell = worksheet.cell(row=row_number, column=column_number)
            if review_row.status == REVIEW_OK:
                cell.fill = _OK_FILL
            elif review_row.status == REVIEW_NEEDS_REVIEW:
                cell.fill = _NEEDS_REVIEW_FILL
                cell.comment = Comment(
                    _cell_comment_text(review_row),
                    _COMMENT_AUTHOR,
                )
            elif review_row.status == REVIEW_NOT_DISCLOSED:
                cell.fill = _NOT_DISCLOSED_FILL
                cell.comment = Comment(
                    _cell_comment_text(review_row),
                    _COMMENT_AUTHOR,
                )


def _client_sheet(workbook: Any) -> Any | None:
    return next(
        (
            sheet
            for sheet in workbook.worksheets
            if sheet.title.startswith(_CLIENT_SHEET_PREFIX)
            or sheet.title == _EXTRACTION_DRAFT_SHEET
        ),
        None,
    )


def _review_rows_by_source(
    review_queue: list[ReviewQueueRow],
) -> dict[str, list[ReviewQueueRow]]:
    rows_by_source: dict[str, list[ReviewQueueRow]] = {}
    for row in review_queue:
        rows_by_source.setdefault(row.source_file, []).append(row)
    return rows_by_source


def _cell_comment_text(row: ReviewQueueRow) -> str:
    reason = row.reason or row.status
    quote = row.source_quote or "No source quote."
    return (
        f"{row.status}: {reason}\n"
        f"Source page: {row.source_page}\n"
        f"Quote: {quote}"
    )


def _write_status_only_workbook(
    out_xlsx: Path, statuses: list[FileStatus]
) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Batch Status"
    worksheet.append(_STATUS_HEADERS)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
    for status in statuses:
        worksheet.append(_status_row(status))
    workbook.save(out_xlsx)
