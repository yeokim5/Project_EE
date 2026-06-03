"""Reviewed export artifacts for the earnings extraction workflow.

This is the gated final stage. A draft only becomes a client workbook after the
export gate passes: every required template field must have a human review
decision, and attention-flagged fields must carry a reviewer note. Demo/synthetic
decisions never produce a "final" workbook -- they require the explicit
``--allow-unreviewed`` escape hatch, which clearly marks the output as a draft.
This is what makes the output trustworthy: numbers reach the client sheet only
because a human approved them, not because the model was confident.

Cell formatting matches the client template exactly: currency fields render as
``$<n>B`` (values are stored in USD millions and divided by 1000 for display),
gross margin renders as a real percentage cell, EPS is a plain number, and any
value that was not approved (or is unsupported for the document type) is left
blank rather than guessed.
"""

from __future__ import annotations

import json
import re
from collections import defaultdict
from copy import copy
from pathlib import Path
from typing import Any, Literal

from openpyxl import load_workbook
from openpyxl.comments import Comment
from pydantic import BaseModel, ConfigDict

from earnings_extractor.capital_return import (
    is_bare_number,
    narrative_from_quote,
    narrative_looks_mangled,
)
from earnings_extractor.review import (
    ReviewDecision,
    ReviewDecisionsFile,
    ReviewItem,
    build_review_items,
    load_review_decisions,
)
from earnings_extractor.schema import TEMPLATE_FIELDS, DraftRun

TEMPLATE_PATH = Path("assesment_info") / "EarningsSample (1).xlsx"
CURRENCY_TEMPLATE_FIELDS = {
    "Total revenue",
    "Net income",
    "Operating income",
    "Operating expenses",
}
GROSS_MARGIN_FIELD = "Gross margin"
EPS_FIELD = "Earnings per share"
CAPITAL_RETURN_FIELD = "Buybacks and dividends"
NON_POPULATING_STATUSES = {"rejected", "needs_fix", "not_applicable"}
ExportDisplayMode = Literal["template", "batch"]


class ExportArtifacts(BaseModel):
    model_config = ConfigDict(extra="forbid")

    xlsx_path: Path
    json_path: Path
    audit_report_path: Path
    is_draft_unreviewed: bool


class ExportGateResult(BaseModel):
    model_config = ConfigDict(extra="forbid")

    allowed: bool
    is_draft_unreviewed: bool
    blocking_reasons: list[str]
    warnings: list[str]


class ExportContext(BaseModel):
    model_config = ConfigDict(extra="forbid", arbitrary_types_allowed=True)

    draft: DraftRun
    review_items: list[ReviewItem]
    decisions: ReviewDecisionsFile
    decisions_by_metric_id: dict[str, ReviewDecision]
    gate: ExportGateResult
    client_rows: list[dict[str, Any]]
    metrics_rows: list[dict[str, Any]]
    decisions_rows: list[dict[str, Any]]
    evidence_rows: list[dict[str, Any]]
    duplicate_warnings: list[str]


def export_reviewed_run(
    run_dir: Path,
    decisions_path: Path,
    out_path: Path,
    allow_unreviewed: bool = False,
    display_mode: ExportDisplayMode = "template",
) -> ExportArtifacts:
    draft = _load_draft(run_dir / "draft_metrics.json")
    review_items = build_review_items(draft)
    decisions = load_review_decisions(decisions_path, expected_run_id=draft.run_id)
    decisions_by_metric_id = {
        decision.metric_id: decision for decision in decisions.decisions
    }
    client_rows, duplicate_warnings = _build_client_rows(
        review_items,
        decisions_by_metric_id,
        allow_unreviewed,
        display_mode,
    )
    gate = _evaluate_export_gate(
        review_items=review_items,
        decisions=decisions,
        decisions_by_metric_id=decisions_by_metric_id,
        allow_unreviewed=allow_unreviewed,
    )
    gate.warnings.extend(duplicate_warnings)
    if not gate.allowed:
        raise ValueError("Export blocked:\n- " + "\n- ".join(gate.blocking_reasons))

    context = ExportContext(
        draft=draft,
        review_items=review_items,
        decisions=decisions,
        decisions_by_metric_id=decisions_by_metric_id,
        gate=gate,
        client_rows=client_rows,
        metrics_rows=_metrics_rows(review_items, decisions_by_metric_id),
        decisions_rows=_decisions_rows(decisions),
        evidence_rows=_evidence_rows(review_items, decisions_by_metric_id),
        duplicate_warnings=duplicate_warnings,
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    json_path = out_path.with_suffix(".json")
    audit_report_path = out_path.with_suffix(".audit.md")
    _write_workbook(context, out_path)
    json_path.write_text(
        json.dumps(_json_payload(context), indent=2) + "\n",
        encoding="utf-8",
    )
    audit_report_path.write_text(_audit_report(context), encoding="utf-8")
    return ExportArtifacts(
        xlsx_path=out_path,
        json_path=json_path,
        audit_report_path=audit_report_path,
        is_draft_unreviewed=gate.is_draft_unreviewed,
    )


def format_currency_billions(value: float | int) -> str:
    billions = float(value) / 1000.0
    rendered = f"{billions:.1f}".rstrip("0").rstrip(".")
    return f"${rendered}B"


def format_currency_millions(value: float | int) -> str:
    return f"${float(value):,.0f} million"


def format_currency_batch_text(value: float | int, source_quote: str) -> str:
    if re.search(r"\bbillions?\b", source_quote, flags=re.IGNORECASE):
        billions = float(value) / 1000.0
        rendered = f"{billions:.1f}".rstrip("0").rstrip(".")
        return f"${rendered} billion"
    return format_currency_millions(value)


def format_gross_margin_cell(value: float | int) -> float:
    return float(value) / 100.0


def map_metric_to_client_cell(
    item: ReviewItem,
    display_mode: ExportDisplayMode = "template",
) -> Any:
    if item.value in (None, ""):
        return ""
    if item.metric_name in CURRENCY_TEMPLATE_FIELDS:
        _require_numeric(item)
        if item.unit != "USD" or item.scale != "millions":
            raise ValueError(
                f"{item.metric_name} must be USD millions before export; got "
                f"unit={item.unit!r}, scale={item.scale!r}."
            )
        if display_mode == "batch":
            return format_currency_batch_text(float(item.value), item.source_quote)
        return format_currency_billions(float(item.value))
    if item.metric_name == GROSS_MARGIN_FIELD:
        _require_numeric(item)
        if item.unit != "percentage points":
            raise ValueError(
                "Gross margin must be percentage points before export; got "
                f"unit={item.unit!r}."
            )
        if display_mode == "batch":
            return f"{float(item.value):g}%"
        return format_gross_margin_cell(float(item.value))
    if item.metric_name == EPS_FIELD:
        _require_numeric(item)
        if display_mode == "batch":
            suffix = (
                " diluted"
                if re.search(r"\bdiluted\b", item.source_quote, flags=re.IGNORECASE)
                else ""
            )
            return f"${float(item.value):.2f}{suffix}"
        return float(item.value)
    if display_mode == "batch" and item.metric_name == "Quarter":
        return _compact_quarter(str(item.value))
    if item.metric_name == CAPITAL_RETURN_FIELD:
        # A bare number ("2100.0") is unconverted raw data, never a verifiable
        # sentence -- but the source quote usually still names the real amount,
        # so read the narrative from there rather than dropping the disclosure.
        text = _capital_return_display_text(item, display_mode)
        if text is None:
            return _missing_client_value(CAPITAL_RETURN_FIELD, display_mode)
        return text
    return str(item.value)


def _load_draft(path: Path) -> DraftRun:
    try:
        return DraftRun.model_validate_json(path.read_text(encoding="utf-8"))
    except FileNotFoundError:
        raise FileNotFoundError(f"Missing draft metrics file: {path}") from None


def _evaluate_export_gate(
    review_items: list[ReviewItem],
    decisions: ReviewDecisionsFile,
    decisions_by_metric_id: dict[str, ReviewDecision],
    allow_unreviewed: bool,
) -> ExportGateResult:
    blockers: list[str] = []
    if decisions.is_demo:
        blockers.append(
            "Review decisions are demo-generated; pass --allow-unreviewed for a "
            "draft/unreviewed export."
        )
    for source_file, items in _group_by_source_file(review_items).items():
        by_name = _items_by_metric_name(items)
        for field in TEMPLATE_FIELDS:
            candidates = by_name.get(field, [])
            label = f"{source_file} / {field}"
            if not candidates:
                blockers.append(
                    f"No metric was extracted for required field {label}."
                )
                continue
            for item in candidates:
                decision = decisions_by_metric_id.get(item.metric_id)
                if decision is None:
                    blockers.append(f"Missing review decision for {label}.")
                    continue
                if item.requires_attention and not _has_note(decision):
                    blockers.append(
                        f"Attention-flagged field {label} requires a reviewer note."
                    )
    if allow_unreviewed:
        return ExportGateResult(
            allowed=True,
            is_draft_unreviewed=True,
            blocking_reasons=[],
            warnings=blockers,
        )
    return ExportGateResult(
        allowed=not blockers,
        is_draft_unreviewed=False,
        blocking_reasons=blockers,
        warnings=[],
    )


def _build_client_rows(
    review_items: list[ReviewItem],
    decisions_by_metric_id: dict[str, ReviewDecision],
    allow_unreviewed: bool,
    display_mode: ExportDisplayMode,
) -> tuple[list[dict[str, Any]], list[str]]:
    rows: list[dict[str, Any]] = []
    warnings: list[str] = []
    for source_file, items in _group_by_source_file(review_items).items():
        by_name = _items_by_metric_name(items)
        row: dict[str, Any] = {"_source_file": source_file}
        for field in TEMPLATE_FIELDS:
            candidates = by_name.get(field, [])
            chosen = _choose_template_item(
                field,
                candidates,
                decisions_by_metric_id,
                allow_unreviewed,
            )
            if len(candidates) > 1:
                warnings.append(
                    f"{source_file} has {len(candidates)} candidates for {field}; "
                    f"used metric_id {chosen.metric_id if chosen else 'none'}."
                )
            decision = (
                decisions_by_metric_id.get(chosen.metric_id) if chosen else None
            )
            if field == CAPITAL_RETURN_FIELD:
                # Narrative field: the quote, not the raw value, is the truth.
                row[field] = _resolve_capital_return_cell(
                    chosen, decision, allow_unreviewed, display_mode
                )
            elif chosen is None or not _should_populate_cell(
                chosen, decision, allow_unreviewed
            ):
                row[field] = _missing_client_value(field, display_mode)
            else:
                row[field] = map_metric_to_client_cell(chosen, display_mode)
        rows.append(row)
    return rows, warnings


def _resolve_capital_return_cell(
    chosen: ReviewItem | None,
    decision: ReviewDecision | None,
    allow_unreviewed: bool,
    display_mode: ExportDisplayMode,
) -> Any:
    """Buybacks cell: derive narrative from the quote, then apply the trust gate.

    Keeps the same approval discipline as every other cell -- an unreviewed
    figure only appears in a draft (``allow_unreviewed``) or once approved -- but
    gates on the derived sentence instead of the raw value, which is routinely
    ``None`` for this narrative field even when the quote names a real amount.
    """

    text = _capital_return_display_text(chosen, display_mode) if chosen else None
    if text is None:
        return _missing_client_value(CAPITAL_RETURN_FIELD, display_mode)
    approved = (
        decision.review_status == "approved"
        if decision is not None
        else allow_unreviewed
    )
    if not approved:
        return _missing_client_value(CAPITAL_RETURN_FIELD, display_mode)
    return text


def _missing_client_value(field: str, display_mode: ExportDisplayMode) -> str:
    if display_mode != "batch":
        return ""
    if field == "Buybacks and dividends":
        return "Not disclosed in this release"
    return "Not disclosed"


def _compact_quarter(value: str) -> str:
    value = value.strip()
    # "First Quarter 2025" / "First-quarter 2025".
    match = re.fullmatch(
        r"(First|Second|Third|Fourth)[\s-]+Quarter\s+(\d{4})",
        value,
        flags=re.IGNORECASE,
    )
    if match:
        quarter = {
            "first": "Q1",
            "second": "Q2",
            "third": "Q3",
            "fourth": "Q4",
        }[match.group(1).lower()]
        return f"{quarter} {match.group(2)}"
    # Compact ledger forms: "1Q25", "1Q2025", "Q1 25", "Q1-2026" -> "Q1 2025".
    compact = re.fullmatch(
        r"Q?([1-4])[\s-]*Q?[\s-]*(\d{2}|\d{4})",
        value,
        flags=re.IGNORECASE,
    )
    if compact:
        quarter, year = compact.group(1), compact.group(2)
        if len(year) == 2:
            year = f"20{year}"
        return f"Q{quarter} {year}"
    # Period-ending phrasing: "Quarter ended Mar 31, 2026",
    # "Three Months Ended March 31, 2026" -> the quarter whose fiscal period
    # closes in that month (only the four quarter-end months map cleanly).
    period = re.search(
        r"(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*\.?\s+\d{1,2},?\s+(\d{4})",
        value,
        flags=re.IGNORECASE,
    )
    if period:
        quarter = {"mar": "Q1", "jun": "Q2", "sep": "Q3", "dec": "Q4"}.get(
            period.group(1).lower()
        )
        if quarter:
            return f"{quarter} {period.group(2)}"
    return value


def _capital_return_display_text(
    item: ReviewItem, display_mode: ExportDisplayMode
) -> str | None:
    """The buybacks cell text: existing narrative, else derived from the quote.

    By export time any live model narrative is already stored on ``item.value``;
    this only has to clean an existing sentence or fall back to the deterministic
    quote deriver (recorded runs and live model misses).
    """

    if (
        isinstance(item.value, str)
        and not is_bare_number(item.value)
        and not narrative_looks_mangled(item.value)
    ):
        if display_mode == "batch":
            return _clean_capital_return_text(item.value)
        return str(item.value)
    return narrative_from_quote(item.source_quote)


def _clean_capital_return_text(value: str) -> str:
    value = re.sub(r"\bworth of share repurchases\b", "of repurchases", value)
    value = re.sub(r"\bto shareholders\b", "", value)
    value = re.sub(r"\bincreased to\b", "of", value)
    value = re.sub(
        r"\sand\s+(\$\d+(?:\.\d+)?)\s+quarterly cash dividend per share\b",
        r"; \1 per share cash dividend",
        value,
    )
    value = re.sub(r"^\$(\d+)\s+billion returned", r"$\1.0 billion returned", value)
    return re.sub(r"\s+", " ", value).strip()


def _choose_template_item(
    field: str,
    candidates: list[ReviewItem],
    decisions_by_metric_id: dict[str, ReviewDecision],
    allow_unreviewed: bool,
) -> ReviewItem | None:
    if not candidates:
        return None
    return sorted(
        candidates,
        key=lambda item: (
            _decision_rank(
                item,
                decisions_by_metric_id.get(item.metric_id),
                allow_unreviewed,
            ),
            item.confidence,
            -item.metric_index,
        ),
        reverse=True,
    )[0]


def _decision_rank(
    item: ReviewItem,
    decision: ReviewDecision | None,
    allow_unreviewed: bool,
) -> int:
    if _should_populate_cell(item, decision, allow_unreviewed):
        return 3
    if decision is not None and decision.review_status in NON_POPULATING_STATUSES:
        return 2
    if decision is not None:
        return 1
    return 0


def _should_populate_cell(
    item: ReviewItem,
    decision: ReviewDecision | None,
    allow_unreviewed: bool,
) -> bool:
    """Decide whether a value is allowed onto the client sheet.

    The gate is intentionally strict: a cell is populated only when a human
    approved it. Missing or non-"approved" decisions leave the cell blank unless
    ``allow_unreviewed`` is set (draft export). A blank cell is the safe default
    -- we never want an unreviewed model guess to look like a confirmed figure.
    """

    if item.value in (None, ""):
        return False
    if decision is None:
        return allow_unreviewed
    return decision.review_status == "approved"


def _group_by_source_file(
    review_items: list[ReviewItem],
) -> dict[str, list[ReviewItem]]:
    grouped: dict[str, list[ReviewItem]] = {}
    for item in review_items:
        grouped.setdefault(item.source_file, []).append(item)
    return grouped


def _items_by_metric_name(items: list[ReviewItem]) -> dict[str, list[ReviewItem]]:
    by_name: dict[str, list[ReviewItem]] = defaultdict(list)
    for item in items:
        if item.metric_name in TEMPLATE_FIELDS:
            by_name[item.metric_name].append(item)
    return by_name


def _write_workbook(context: ExportContext, out_path: Path) -> None:
    workbook = load_workbook(TEMPLATE_PATH)
    worksheet = workbook.worksheets[0]
    worksheet.title = (
        "Client Template DRAFT"
        if context.gate.is_draft_unreviewed
        else "Client Template"
    )
    if context.gate.is_draft_unreviewed:
        worksheet.sheet_properties.tabColor = "FFC000"
        worksheet["A1"].comment = Comment(
            "DRAFT/UNREVIEWED export generated with --allow-unreviewed.",
            "earnings_extractor",
        )

    sample_styles = [
        _copy_cell_style(worksheet.cell(row=2, column=col))
        for col in range(1, 10)
    ]
    if worksheet.max_row > 1:
        worksheet.delete_rows(2, worksheet.max_row - 1)
    for row_number, client_row in enumerate(context.client_rows, start=2):
        for column_number, field in enumerate(TEMPLATE_FIELDS, start=1):
            cell = worksheet.cell(row=row_number, column=column_number)
            _apply_cell_style(cell, sample_styles[column_number - 1])
            cell.value = client_row[field]

    _add_tab(workbook, "Metrics", context.metrics_rows)
    _add_tab(workbook, "Review Decisions", context.decisions_rows)
    _add_tab(workbook, "Evidence", context.evidence_rows)
    workbook.save(out_path)


def _copy_cell_style(cell: Any) -> dict[str, Any]:
    return {
        "font": copy(cell.font),
        "fill": copy(cell.fill),
        "border": copy(cell.border),
        "alignment": copy(cell.alignment),
        "number_format": cell.number_format,
        "protection": copy(cell.protection),
    }


def _apply_cell_style(cell: Any, style: dict[str, Any]) -> None:
    cell.font = copy(style["font"])
    cell.fill = copy(style["fill"])
    cell.border = copy(style["border"])
    cell.alignment = copy(style["alignment"])
    cell.number_format = style["number_format"]
    cell.protection = copy(style["protection"])


def _add_tab(workbook: Any, title: str, rows: list[dict[str, Any]]) -> None:
    if title in workbook.sheetnames:
        del workbook[title]
    worksheet = workbook.create_sheet(title)
    headers = list(rows[0].keys()) if rows else []
    worksheet.append(headers)
    for row in rows:
        worksheet.append([row.get(header) for header in headers])


def _metrics_rows(
    review_items: list[ReviewItem],
    decisions_by_metric_id: dict[str, ReviewDecision],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for item in review_items:
        decision = decisions_by_metric_id.get(item.metric_id)
        row = item.model_dump(mode="json")
        row.update(_decision_summary(decision))
        rows.append(row)
    return rows


def _decisions_rows(decisions: ReviewDecisionsFile) -> list[dict[str, Any]]:
    return [
        {
            "run_id": decisions.run_id,
            "is_demo": decisions.is_demo,
            **decision.model_dump(mode="json"),
        }
        for decision in decisions.decisions
    ]


def _evidence_rows(
    review_items: list[ReviewItem],
    decisions_by_metric_id: dict[str, ReviewDecision],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for item in review_items:
        decision = decisions_by_metric_id.get(item.metric_id)
        rows.append(
            {
                "metric_id": item.metric_id,
                "company": item.company,
                "metric_name": item.metric_name,
                "source_file": item.source_file,
                "source_page": item.source_page,
                "source_quote": item.source_quote,
                "confidence": item.confidence,
                "needs_review": item.needs_review,
                "review_reason": item.review_reason,
                "decision_status": decision.review_status if decision else None,
            }
        )
    return rows


def _decision_summary(decision: ReviewDecision | None) -> dict[str, Any]:
    if decision is None:
        return {
            "decision_status": None,
            "reviewer": None,
            "reviewed_at": None,
            "decision_reviewer_note": None,
        }
    return {
        "decision_status": decision.review_status,
        "reviewer": decision.reviewer,
        "reviewed_at": decision.reviewed_at,
        "decision_reviewer_note": decision.reviewer_note,
    }


def _json_payload(context: ExportContext) -> dict[str, Any]:
    return {
        "run_id": context.draft.run_id,
        "draft_created_at": context.draft.created_at,
        "is_draft_unreviewed": context.gate.is_draft_unreviewed,
        "decisions_is_demo": context.decisions.is_demo,
        "client_rows": context.client_rows,
        "metrics": context.metrics_rows,
        "review_decisions": context.decisions_rows,
        "evidence": context.evidence_rows,
        "warnings": context.gate.warnings,
        "blocking_reasons": context.gate.blocking_reasons,
    }


def _audit_report(context: ExportContext) -> str:
    populated = sum(
        1
        for row in context.client_rows
        for field in TEMPLATE_FIELDS
        if row.get(field) not in (None, "")
    )
    blanks = len(context.client_rows) * len(TEMPLATE_FIELDS) - populated
    export_state = "draft/unreviewed" if context.gate.is_draft_unreviewed else "final"
    lines = [
        "# Export Audit Report",
        "",
        f"- Run ID: `{context.draft.run_id}`",
        f"- Export state: `{export_state}`",
        f"- Decisions demo file: `{context.decisions.is_demo}`",
        f"- Source documents: `{len(context.draft.documents)}`",
        f"- Populated client fields: `{populated}`",
        f"- Blank client fields: `{blanks}`",
        "",
    ]
    if context.gate.warnings:
        lines.extend(["## Warnings", ""])
        lines.extend(f"- {warning}" for warning in context.gate.warnings)
        lines.append("")
    lines.extend(["## Evidence Summary", ""])
    for row in context.evidence_rows:
        lines.append(
            "- "
            f"{row['company'] or 'Unknown'} / {row['metric_name']} "
            f"({row['source_file']} p. {row['source_page']}): "
            f"{row['source_quote']}"
        )
    lines.append("")
    return "\n".join(lines)


def _has_note(decision: ReviewDecision) -> bool:
    return bool((decision.reviewer_note or "").strip())


def _require_numeric(item: ReviewItem) -> None:
    if not isinstance(item.value, int | float):
        raise ValueError(f"{item.metric_name} must be numeric before export.")
